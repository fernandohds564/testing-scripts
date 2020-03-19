#!/usr/bin/env python-sirius

from openpyxl import load_workbook
from copy import deepcopy as _dcopy
import matplotlib.pyplot as plt

fname = '/home/fernando/Download/Booster_Rotation_Deviations.xlsx'
wb = load_workbook(fname, data_only=True)

ws = wb['Planilha1']

tradutor = {
    'QF': 'Quadrupolo',
    'QD': 'QB Dipolo',
    'SF': 'Sextupolo',
    'SD': 'SB Dipolo',
    'B': 'Dipolo'}

err_tmplt = {k: [] for k in tradutor}

err_types = {
    'x': (6, 1, 1000),
    'y': (6, 2, 1000),
    'z': (6, 3, 1000),
    'roll': (11, 2, 1000*3.141592653589793/180)}
errs = {err: _dcopy(err_tmplt) for err in err_types}

for i, row in enumerate(ws.iter_rows(), 1):
    val = row[0].value
    if not isinstance(val, str):
        continue
    for k, v in tradutor.items():
        if not val.startswith(v):
            continue
        for err, consts in err_types.items():
            line, col, conv = consts
            errs[err][k].append(ws[i+line][col].value * conv)
        break

#plt.plot(errs['x']['QF'])
#plt.plot(errs['x']['B'])
#plt.show()
